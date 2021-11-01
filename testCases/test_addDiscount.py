import time

import allure
import openpyxl
import pytest
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager

from pageObjects.AddDiscount import AddDiscount
from pageObjects.LoginPage import LoginPage
from utilities.customLogger import LogGen
from utilities.readProperties import ReadConfig


def readData():
    list = []
    # path = "C:\\Users\\Arumugam\\PycharmProjects\\ECommerce_Demo\\InputData.xlsx"
    path = ".//TestData/InputData.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook["discount"]

    rows = sheet.max_row
    cols = sheet.max_column

    print("Rows->", rows)
    print("Columns", cols)

    for r in range(2, rows + 1):
        discount_name = sheet.cell(r, 1).value
        discount_type = sheet.cell(r, 2).value
        discount_amt = sheet.cell(r, 3).value
        start_date = sheet.cell(r, 4).value
        end_date = sheet.cell(r, 5).value
        discount_limit = sheet.cell(r, 6).value
        search_discount_type = sheet.cell(r, 7).value
        search_discount_name = sheet.cell(r, 8).value

    tuple = (discount_name, discount_type, discount_amt, start_date, end_date, discount_limit, search_discount_type,search_discount_name)
    list.append(tuple)
    return list


class Test_003_AddDiscount:
    baseURL = ReadConfig.getApplicationURL()
    username = ReadConfig.getUsername()
    password = ReadConfig.getPassword()
    logger = LogGen.loggen()

    @pytest.fixture()
    def test_setup(self):
        global driver
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.driver.implicitly_wait(10)
        self.driver.maximize_window()
        yield
        self.driver.quit()

    @allure.description("**********Add customer details information**************")
    @allure.severity(severity_level="NORMAL")
    @pytest.mark.parametrize(
        "discount_name,discount_type,discount_amt,start_date,end_date,discount_limit,search_discount_type,search_discount_name",
        readData())
    def test_addDiscount(self, test_setup, discount_name, discount_type, discount_amt, start_date, end_date, discount_limit,
                         search_discount_type, search_discount_name):
        self.logger.info("***********Test_003_Add_New_Discount details*************")

        #self.driver = setup
        self.driver.get(self.baseURL)
        self.driver.maximize_window()
        self.lp = LoginPage(self.driver)
        self.lp.setUserName(self.username)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.logger.info("***********Login Successful*************")
        self.discount = AddDiscount(self.driver)
        self.discount.clickPromotionLink()
        self.discount.clickDiscountLink()
        self.discount.clickAddNew()
        print('enter discount name')
        self.discount.enterDiscountName(discount_name)
        self.discount.selectDiscountType(discount_type)
        self.discount.enterDiscountAmount(discount_amt)
        self.discount.enterStartDate(start_date)
        self.discount.enterEndDate(end_date)
        self.discount.selectDiscountLimit(discount_limit)
        time.sleep(5)
        self.discount.clickSaveBtn()
        self.logger.info("Discount details entered successfully")
        self.logger.info("Search Discount details and edit information")
        self.discount.searchDiscountType(search_discount_type)
        self.discount.enterSearchDiscountName(search_discount_name)
        self.discount.clickSearchBtn()
        self.driver.close()
        self.logger.info("******driver closed***********")
