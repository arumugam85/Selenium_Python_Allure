import time
import allure
import openpyxl
import pytest
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager

from pageObjects.AddProducts import AddProducts
from pageObjects.LoginPage import LoginPage
from utilities.customLogger import LogGen
from utilities.readProperties import ReadConfig

def readData():
    list = []
    # path = "C:\\Users\\Arumugam\\PycharmProjects\\ECommerce_Demo\\InputData.xlsx"
    path = ".//TestData/InputData.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook["products"]

    rows = sheet.max_row
    cols = sheet.max_column

    for r in range(2, rows + 1):
        prod_name = sheet.cell(r, 1).value
        prod_desc = sheet.cell(r, 2).value
        sku_name = sheet.cell(r, 3).value
        category_name = sheet.cell(r, 4).value
        mfg_name = sheet.cell(r, 5).value
        prod_tagname = sheet.cell(r, 6).value
        gtin_number = sheet.cell(r, 7).value
        mfg_number = sheet.cell(r, 8).value
        prod_type = sheet.cell(r, 9).value
        prod_template = sheet.cell(r, 10).value
        cust_role = sheet.cell(r, 11).value
        start_date = sheet.cell(r, 12).value
        end_date = sheet.cell(r, 13).value

        tuple = (
        prod_name, prod_desc, sku_name, category_name, mfg_name, prod_tagname, gtin_number, mfg_number, prod_type,
        prod_template, cust_role, start_date, end_date)
        list.append(tuple)
        return list


class Test_003_AddProducts:
    baseURL = ReadConfig.getApplicationURL()
    username = ReadConfig.getUsername()
    password = ReadConfig.getPassword()
    logger = LogGen.loggen()

    @pytest.fixture()
    def test_setup(self):
        global driver
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.driver.implicitly_wait(10)
        self.driver.maximize_window()
        yield
        self.driver.quit()

    @allure.description("**********Add customer details information**************")
    @allure.severity(severity_level="NORMAL")
    @pytest.mark.parametrize(
        "prod_name,prod_desc,sku_name,category_name,mfg_name,prod_tagname,gtin_number,mfg_number,prod_type,prod_template,cust_role,start_date,end_date",
        readData())
    def test_addProducts(self, test_setup, prod_name, prod_desc, sku_name, category_name, mfg_name, prod_tagname,
                         gtin_number, mfg_number, prod_type, prod_template, cust_role, start_date, end_date):
        self.logger.info("***********Test_003_Add_New_Product details*************")

       # self.driver = setup
        self.driver.get(self.baseURL)
        self.driver.maximize_window()
        self.lp = LoginPage(self.driver)
        self.lp.setUserName(self.username)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.logger.info("***********Login Successful*************")
        self.products = AddProducts(self.driver)
        self.products.clickCategoryLink()
        self.products.clickProductsLink()
        self.products.clickAddNewBtn()
        self.products.enterProductName(prod_name)
        self.products.enterProductDesc(prod_desc)
        self.products.enterSKUName(sku_name)

        #self.products.clickSaveBtn()
        # self.products.selectManufacturingName(mfg_name)
        self.products.enterProductTagName(prod_tagname)
        self.products.enterGTINNumber(gtin_number)
        self.products.enterMFGNumber(mfg_number)
        self.products.selectProdType(prod_type)
        self.products.enterProdTemplate(prod_template)
        self.products.selectCustRoles(cust_role)
        #self.products.clickSaveBtn()
        self.products.enterStartDate(start_date)
        self.products.enterEndDate(end_date)
        self.products.selectCategories(category_name)
        time.sleep(5)
        self.products.clickSaveBtn()
        self.logger.info("Product details added successfully")
        self.driver.close()
        self.logger.info("******driver closed***********")
