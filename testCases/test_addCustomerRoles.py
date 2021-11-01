import allure
import pytest
import openpyxl
import self
from allure_commons.types import AttachmentType
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager

from pageObjects.AddCustomerPage import AddCustomer
from pageObjects.AddCustomerRole import AddCustomerRole
from pageObjects.LoginPage import LoginPage
from testCases.conftest import BaseClass
from utilities.XLUtils import readData
from utilities.customLogger import LogGen
from utilities.readProperties import ReadConfig


def readData():
    list = []
    # path = "C:\\Users\\Arumugam\\PycharmProjects\\ECommerce_Demo\\InputData.xlsx"
    path = ".//TestData/InputData.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook["addcustrole"]

    rows = sheet.max_row
    cols = sheet.max_column

    for r in range(2, rows + 1):
        # custname = sheet.cell(r, 1).value
        # showcount = sheet.cell(r, 2).value
        # firstname = sheet.cell(r, 3).value
        # lastname = sheet.cell(r, 4).value
        # gender = sheet.cell(r, 5).value
        # dob = sheet.cell(r, 6).value
        # company = sheet.cell(r, 7).value
        # newsletter = sheet.cell(r, 8).value
        # role = sheet.cell(r, 9).value
        # vendor = sheet.cell(r, 10).value
        # comment = sheet.cell(r, 11).value

        # tuple = (custname)
        # tuple = (custname, custpassword, firstname, lastname, gender, dob, company, newsletter, role, vendor, comment)
        list.append(tuple)
        return list


@pytest.mark.usefixtures('test_setup')
class Test_006_AddCustomerRole:
    baseURL = ReadConfig.getApplicationURL()
    username = ReadConfig.getUsername()
    password = ReadConfig.getPassword()
    logger = LogGen.loggen()

    @pytest.fixture()
    def test_setup(self):
        global driver
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.driver.implicitly_wait(10)
        self.driver.maximize_window()
        yield
        self.driver.quit()

    @allure.description("**********Add customer Role details information**************")
    @allure.severity(severity_level="NORMAL")
    # @pytest.mark.parametrize(readData())
    def test_addCustomerRoles(self, test_setup):
        self.logger.info("***********Test_002_Add_New_Customer*************")
        #self.driver = test_setup
        # self.base = BaseClass(self.driver)
        # self.base.test_setup()
        self.driver.get(self.baseURL)
        self.driver.maximize_window()
        self.lp = LoginPage(self.driver)
        self.lp.setUserName(self.username)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        # ss.Screenshot(ss_path + "login.png")
        self.logger.info("******Login Successful***********")

        self.logger.info("******Add New Customer Roles*********")
        self.addcust = AddCustomer(self.driver)
        self.addcust.clickCustomerLink()
        self.addRole = AddCustomerRole(self.driver)
        # self.addRole.clickCustomerLink()
        print("click customer role link")
        self.addRole.clickCustomerRole()
        # self.addRole.clickAddNewBtn()
        #
        # self.logger.info("******Add Customer Role Information*********")
        # # self.email=random_generator()+"@gmail.com"
        # self.addRole.enterCustomerName(custname)
        # print(custname)
        # self.addRole.clickCustomerActiveChkBox()
        # self.addRole.clickCustomerFreeShipChkBox()
        # self.addRole.clickCustomerTaxChkBox()
        # self.addRole.clickCustomerTaxDisplayChkBox()
        # self.addRole.clickCustomerEnablePass()
        # self.addRole.clickCustomerPurchaseBtn()
        # self.addRole.clickSaveButton()
        self.logger.info("******Add Customer Role Edit Information*********")
        self.addRole.clickEditBtnInWebTable()
        self.addRole.clickCustomerFreeShipChkBox()
        # self.addRole.clickCustomerTaxChkBox()
        self.addRole.clickSaveButton()
        print('Customer role has been edited sucessfully')

        self.logger.info("******Add Customer Role Delete Information*********")
        self.addRole.clickEditButton()
        self.addRole.clickDeleteButton()
        self.addRole.clickConfirmDeleteButton()
        self.msg = self.driver.find_element_by_tag_name("body").text
        print(self.msg)

        if 'The customer role has been deleted successfully.' in self.msg:
            assert True == True
            self.logger.info("**** Add Customer Role test case is passed  ****")
            allure.attach(self.driver.get_screenshot_as_png(), name="addcustomerscreen",
                          attachment_type=AttachmentType.PNG)

        else:
            assert True == False
            self.logger.info("**** Add Customer Role test case is failed  ****")

            allure.attach(self.driver_get_screenshot_as_png(), name="addcustomerscreen",
                          attachment_type=AttachmentType.PNG)

        self.driver.close()
        self.logger.info("******driver closed***********")

# def random_generator(size=8,chars=string.ascii_lowercase + string.digits):
#     return ''.join(random.choice(chars) for x in range(size))
