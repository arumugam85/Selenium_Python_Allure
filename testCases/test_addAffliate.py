import time
import allure
import openpyxl
import pytest
from allure_commons.types import AttachmentType
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager

from pageObjects.AddAffliate import AddAffliate
from pageObjects.AddDiscount import AddDiscount
from pageObjects.LoginPage import LoginPage
from utilities.customLogger import LogGen
from utilities.readProperties import ReadConfig


def readData():
    list = []
    # path = "C:\\Users\\Arumugam\\PycharmProjects\\ECommerce_Demo\\InputData.xlsx"
    path = ".//TestData/InputData.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook["affliate"]

    rows = sheet.max_row
    cols = sheet.max_column

    for r in range(2, rows + 1):
        first_name = sheet.cell(r, 1).value
        last_name = sheet.cell(r, 2).value
        email = sheet.cell(r, 3).value
        company_name = sheet.cell(r, 4).value
        country_name = sheet.cell(r, 5).value
        state_name = sheet.cell(r, 6).value
        region_name = sheet.cell(r, 7).value
        city_name = sheet.cell(r, 8).value
        address = sheet.cell(r, 9).value
        zip_code = sheet.cell(r, 10).value
        phone_number = sheet.cell(r, 11).value
        fax_number = sheet.cell(r, 12).value
        comp_name = sheet.cell(r, 13).value

        tuple = (
            first_name, last_name, email, company_name, country_name, state_name, region_name, city_name, address,
            zip_code,
            phone_number, fax_number, comp_name)
        list.append(tuple)
        return list


class Test_003_CRUDAffliate:
    baseURL = ReadConfig.getApplicationURL()
    username = ReadConfig.getUsername()
    password = ReadConfig.getPassword()
    logger = LogGen.loggen()

    @pytest.fixture()
    def test_setup(self):
        global driver
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.driver.implicitly_wait(10)
        self.driver.maximize_window()
        yield
        self.driver.quit()

    @allure.description("**********Add Affliation details information**************")
    @allure.severity(severity_level="NORMAL")
    @pytest.mark.parametrize(
        "first_name, last_name, email, company_name, country_name, state_name, region_name, city_name, address, zip_code,phone_number, fax_number,comp_name",
        readData())
    def test_addAffliate(self, test_setup, first_name, last_name, email, company_name, country_name, state_name, region_name,
                         city_name, address, zip_code, phone_number, fax_number, comp_name):
        self.logger.info("***********Test_003_Add_New_Affliation details*************")

        #self.driver = setup
        self.driver.get(self.baseURL)
        self.driver.maximize_window()
        self.logger.info("***********Enter Login Information*************")
        self.lp = LoginPage(self.driver)
        self.lp.setUserName(self.username)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.logger.info("***********Enter Affliation Information*************")
        self.discount = AddDiscount(self.driver)
        self.discount.clickPromotionLink()
        self.affliate = AddAffliate(self.driver)
        self.affliate.clickAffliate()
        self.affliate.clickAddNewBtn()
        self.affliate.enterFirstName(first_name)
        self.affliate.enterLastName(last_name)
        self.affliate.enterEmail(email)
        self.affliate.enterCompanyName(company_name)
        self.affliate.selectCountry(country_name)
        self.affliate.selectState(state_name)
        self.affliate.enterCounty(region_name)
        self.affliate.enterCityName(city_name)
        self.affliate.enterAddress(address)
        self.affliate.enterZipCode(zip_code)
        self.affliate.enterPhoneNumber(phone_number)
        self.affliate.enterFaxNumber(fax_number)
        self.affliate.clickSaveBtn()
        time.sleep(3)
        #self.affliate.verifySuccessMsg()

        self.logger.info("***********Test_003_Edit_New_Affliation details*************")
        self.affliate.clickEditButton()
        time.sleep(3)
        self.affliate.editCompanyName(comp_name)
        self.affliate.clickSaveBtn()
        time.sleep(3)
        # self.affliate.verifySuccessMsg()
        self.logger.info("***********Test_003_Delete_Affliation details*************")
        time.sleep(3)
        self.affliate.clickEditButton()
        time.sleep(3)
        self.affliate.clickDeleteButton()
        time.sleep(3)
        self.logger.info("Affliation details deleted successfully")
        self.driver.close()
        self.logger.info("******driver closed***********")
