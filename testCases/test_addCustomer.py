import allure
import pytest
import openpyxl
from allure_commons.types import AttachmentType

from pageObjects.AddCustomerPage import AddCustomer
from pageObjects.LoginPage import LoginPage
from utilities.XLUtils import readData
from utilities.customLogger import LogGen
from utilities.readProperties import ReadConfig


class Test_002_AddCustomer:
    baseURL = ReadConfig.getApplicationURL()
    username = ReadConfig.getUsername()
    password = ReadConfig.getPassword()
    logger = LogGen.loggen()

    @pytest.fixture()
    def test_setup(self):
        global driver
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.driver.implicitly_wait(10)
        self.driver.maximize_window()
        yield
        self.driver.quit()

    def readData():
        list = []
        # path = "C:\\Users\\Arumugam\\PycharmProjects\\ECommerce_Demo\\InputData.xlsx"
        path = ".//TestData/InputData.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook["addcust"]

        rows = sheet.max_row
        cols = sheet.max_column

        for r in range(2, rows + 1):
            email = sheet.cell(r, 1).value
            custpassword = sheet.cell(r, 2).value
            firstname = sheet.cell(r, 3).value
            lastname = sheet.cell(r, 4).value
            gender = sheet.cell(r, 5).value
            dob = sheet.cell(r, 6).value
            company = sheet.cell(r, 7).value
            newsletter = sheet.cell(r, 8).value
            role = sheet.cell(r, 9).value
            vendor = sheet.cell(r, 10).value
            comment = sheet.cell(r, 11).value

            tuple = (email, custpassword, firstname, lastname, gender, dob, company, newsletter, role, vendor, comment)
            list.append(tuple)
            return list

    @allure.description("**********Add customer details information**************")
    @allure.severity(severity_level="NORMAL")
    @pytest.mark.parametrize("email,custpassword,firstname,lastname,gender,dob,company,newsletter,role,vendor,comment",readData())
    def test_addCustomer(self, test_setup,email,custpassword,firstname,lastname,gender,dob,company,newsletter,role,vendor,comment):
        self.logger.info("***********Test_002_Add_New_Customer*************")
        # ss = ScreenShots(driver)
        # ss_path = "/test_addcustomer/"
        self.driver = setup
        self.driver.get(self.baseURL)
        self.driver.maximize_window()
        self.lp = LoginPage(self.driver)
        self.lp.setUserName(self.username)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        # ss.Screenshot(ss_path + "login.png")
        self.logger.info("******Login Successful***********")

        self.logger.info("******Add New Customer*********")
        self.addcust = AddCustomer(self.driver)
        self.addcust.clickCustomerLink()
        self.addcust.clickCustomerList()
        self.addcust.btnAddNew()

        self.logger.info("******Enter Customer Information*********")
        #self.email=random_generator()+"@gmail.com"
        self.addcust.enterEmail(email)
        print(email)
        self.addcust.enterPassword(custpassword)
        self.logger.info("******Enter firstname*********")
        self.addcust.enterFirstName(firstname)
        self.logger.info("******Enter lastename*********")
        self.addcust.enterLastName(lastname)
        self.addcust.clickGender(gender)
        self.addcust.enterDob(dob)
        self.addcust.clickTaxExempt()
        self.addcust.enterCompanyName(company)
        #self.addcust.selectNewsLetterDropdown(newsletter)
        #self.addcust.selectCustomerRole(role)
        self.addcust.selectVendor(vendor)
        self.addcust.enterComments(comment)
        self.addcust.clickSaveBtn()
        print("save btn clicked")

        self.msg =self.driver.find_element_by_tag_name("body").text
        print(self.msg)

        if 'The new customer has been added successfully.' in self.msg:
            assert True == True
            self.logger.info("**** Add Customer test case is passed  ****")
            allure.attach(self.driver.get_screenshot_as_png(), name="addcustomerscreen",
                          attachment_type=AttachmentType.PNG)

        else:
            assert True == False
            self.logger.info("**** Add Customer test case is failed  ****")

            allure.attach(self.driver_get_screenshot_as_png(), name="addcustomerscreen", attachment_type=AttachmentType.PNG)

        self.driver.quit()
        self.logger.info("******driver closed***********")


# def random_generator(size=8,chars=string.ascii_lowercase + string.digits):
#     return ''.join(random.choice(chars) for x in range(size))
